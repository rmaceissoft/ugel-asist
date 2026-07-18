from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Protection

from . import db


ROOT_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = ROOT_DIR / "assets" / "UGEL - Plantilla de asistencia.xlsx"

ANEXO3_START_ROW = 13
ANEXO4_START_ROW = 12
TEACHER_INFO_COLUMNS = (1, 2, 3, 5, 6, 7, 8, 9)


def generate_school_workbook(version_id: int, colegio_codigo: str) -> bytes:
    with db.connect() as conn:
        docentes = db.get_docentes(conn, version_id, colegio_codigo)
        version = db.get_version(conn, version_id)

    if not docentes:
        raise ValueError("No hay docentes para el colegio seleccionado.")

    workbook = load_workbook(TEMPLATE_PATH)
    _fill_sheet(workbook["ANEXO 3"], docentes, version, ANEXO3_START_ROW)
    _fill_sheet(workbook["ANEXO 4"], docentes, version, ANEXO4_START_ROW)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_changed_schools_zip(version_id: int) -> bytes:
    output = BytesIO()
    with db.connect() as conn:
        schools = db.get_changed_schools(conn, version_id, only_changed=True)

    with ZipFile(output, "w") as archive:
        for school in schools:
            content = generate_school_workbook(version_id, school["colegio_codigo"])
            filename = _safe_filename(
                f"{school['colegio_codigo']} - {school['colegio_nombre']}.xlsx"
            )
            archive.writestr(filename, content)
    return output.getvalue()


def _fill_sheet(sheet, docentes, version, start_row: int) -> None:
    first = docentes[0]
    _unlock_sheet_cells(sheet)
    _set_merged_value(sheet, "D6", first["colegio_nombre"])
    _set_merged_value(sheet, "D8", first["nivel_educativo"])
    if version:
        period = version["date_time"][:7]
        # Period cells differ between ANEXO 3 and ANEXO 4 in the template.
        if sheet.title == "ANEXO 3":
            _set_merged_value(sheet, "P4", period)
        else:
            _set_merged_value(sheet, "R3", period)

    _ensure_rows(sheet, start_row, len(docentes))
    for index, docente in enumerate(docentes, start=start_row):
        number = index - start_row + 1
        full_name = _full_name(docente)
        values = {
            1: number,
            2: docente["dni"],
            3: full_name,
            5: docente["cargo"],
            6: docente["especialidad"],
            7: docente["situacion_laboral"],
            8: docente["celular"],
            9: docente["email"],
        }
        for column, value in values.items():
            _set_cell_value(sheet, index, column, value)
        _lock_teacher_info_cells(sheet, index)
    _protect_sheet(sheet)


def _ensure_rows(sheet, start_row: int, needed_rows: int) -> None:
    existing_capacity = max(sheet.max_row - start_row + 1, 0)
    if needed_rows <= existing_capacity:
        return

    rows_to_add = needed_rows - existing_capacity
    append_start = sheet.max_row + 1
    sheet.insert_rows(append_start, rows_to_add)
    template_row = start_row
    for row in range(append_start, append_start + rows_to_add):
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(row=template_row, column=column)
            target = sheet.cell(row=row, column=column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.border:
                target.border = copy(source.border)
            if source.fill:
                target.fill = copy(source.fill)
            if source.font:
                target.font = copy(source.font)


def _set_merged_value(sheet, cell_ref: str, value) -> None:
    cell = sheet[cell_ref]
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            sheet.cell(merged_range.min_row, merged_range.min_col, value=value)
            return
    cell.value = value


def _set_cell_value(sheet, row: int, column: int, value) -> None:
    coordinate = sheet.cell(row=row, column=column).coordinate
    for merged_range in sheet.merged_cells.ranges:
        if coordinate in merged_range:
            sheet.cell(merged_range.min_row, merged_range.min_col, value=value)
            return
    sheet.cell(row=row, column=column, value=value)


def _unlock_sheet_cells(sheet) -> None:
    unlocked = Protection(locked=False)
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = copy(unlocked)


def _lock_teacher_info_cells(sheet, row: int) -> None:
    locked = Protection(locked=True)
    for column in TEACHER_INFO_COLUMNS:
        coordinate = sheet.cell(row=row, column=column).coordinate
        for merged_range in sheet.merged_cells.ranges:
            if coordinate in merged_range:
                sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).protection = copy(locked)
                break
        else:
            sheet.cell(row=row, column=column).protection = copy(locked)


def _protect_sheet(sheet) -> None:
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True


def _full_name(docente) -> str:
    apellidos = " ".join(
        part
        for part in [docente["primer_apellido"], docente["segundo_apellido"]]
        if part
    )
    if apellidos and docente["nombre"]:
        return f"{apellidos}, {docente['nombre']}"
    return apellidos or docente["nombre"] or ""


def _safe_filename(name: str) -> str:
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if char in invalid else char for char in name)
    return cleaned[:180]
